import scrapy
import openpyxl

class CompanyDetailsSpider(scrapy.Spider):
    name = "company_details"
    allowed_domains = ["europages.co.uk"]

    def start_requests(self):
        yield scrapy.Request(
            url="https://www.europages.co.uk/CEG-ELETTRONICA-INDUSTRIALE/SEAC000061757-001.html", 
            callback=self.parse
        )

    def parse(self, response):
        company_name = response.xpath("//h1[@class='ep-epages-header-title text-h6 text-sm-h4']/text()").get()
        company_address = response.xpath("//p[@class='ma-0']/text()").getall()
        company_country = response.xpath("//span[@class='font-weight-bold']/text()").get()
        company_website = response.xpath("//a[@class='ep-epages-home-link-card v-card v-sheet v-sheet--outlined theme--light pa-4 ep-epages-home-website-link v-card v-card--link v-sheet theme--light']/@href").get()
        company_info = response.xpath("//dd[@class='ep-key-value__value text-body-1']/text()").getall()
        company_specialize = response.xpath("//li[@class='ep-keywords__list-item black black--text body-2 rounded-sm px-2 py-1']/text()").getall()

        company_size = company_info[0] if len(company_info) > 0 else ''
        company_established = company_info[1] if len(company_info) > 1 else ''
        address_street = company_address[0] if len(company_address) > 0 else ''
        address_town = company_address[1] if len(company_address) > 1 else ''

        specializations_list = []
        for each_specialization in company_specialize:
            cleaned_specialization = each_specialization.replace('\n', '').replace('  ', '')
            specializations_list.append(cleaned_specialization)

        company_data = {
            'Platform use': 'https://www.europages.co.uk/',
            'Company Name': company_name.strip() if company_name else '',
            'Address street': address_street.strip(),
            'Address town': address_town.strip(),
            'Country Name': company_country.strip() if company_country else '',
            'Website': company_website.strip() if company_website else '',
            'Established': company_established.strip(),
            'Nr. Of employees': company_size.strip(),
            'Specialize in': ', '.join(specializations_list),
        }

        self.save_to_excel(company_data)

    def save_to_excel(self, data):
        file_path = "company_details.xlsx"

        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            headers = ['Sr. No', 'Platform use', 'Company Name', 'Address street', 'Address town', 'Country Name', 'Website', 'Established', 'Nr. Of employees', 'Specialize in']
            sheet.append(headers)
        else:
            sheet = workbook.active

        # Calculate the next serial number by checking the number of rows
        next_sr_no = sheet.max_row

        row = [
            next_sr_no, data['Platform use'], data['Company Name'], data['Address street'], 
            data['Address town'], data['Country Name'], data['Website'], 
            data['Established'], data['Nr. Of employees'], data['Specialize in']
        ]
        sheet.append(row)

        workbook.save(file_path)
